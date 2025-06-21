import os
import yfinance as yf
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import pickle
from pathlib import Path
import time
warnings.filterwarnings('ignore')

class DateManager:
    """Gestión dinámica e inteligente de fechas"""
    
    @staticmethod
    def get_dynamic_date_ranges():
        """Calcula rangos de fechas dinámicos para análisis"""
        current_date = datetime.now()
        current_year = current_date.year
        
        # Para análisis histórico: años completos
        start_year_10y = current_year - 10  # 2015
        start_year_5y = current_year - 5    # 2020
        end_year_historical = current_year - 1  # 2024
        
        ranges = {
            # Datos históricos completos (años cerrados)
            'historical_10y': (f"{start_year_10y}-01-01", f"{end_year_historical}-12-31"),
            'historical_5y': (f"{start_year_5y}-01-01", f"{end_year_historical}-12-31"),
            
            # Year-to-Date actual
            'ytd': (f"{current_year}-01-01", current_date.strftime("%Y-%m-%d")),
            
            # Rango completo para descarga única
            'full_download': (f"{start_year_10y}-01-01", current_date.strftime("%Y-%m-%d")),
            
            # Metadatos
            'years': {
                'current': current_year,
                'start_10y': start_year_10y,
                'start_5y': start_year_5y,
                'end_historical': end_year_historical
            }
        }
        
        return ranges
    
    @staticmethod
    def get_analysis_periods():
        """Define los períodos de análisis basados en fechas dinámicas"""
        ranges = DateManager.get_dynamic_date_ranges()
        
        return {
            'ytd': {
                'name': f"YTD {ranges['years']['current']}",
                'start_date': ranges['ytd'][0],
                'end_date': ranges['ytd'][1],
                'years_count': 1
            },
            '5y': {
                'name': f"5 Años ({ranges['years']['start_5y']}-{ranges['years']['end_historical']})",
                'start_date': ranges['historical_5y'][0],
                'end_date': ranges['historical_5y'][1],
                'years_count': 5
            },
            '10y': {
                'name': f"10 Años ({ranges['years']['start_10y']}-{ranges['years']['end_historical']})",
                'start_date': ranges['historical_10y'][0],
                'end_date': ranges['historical_10y'][1],
                'years_count': 10
            }
        }

class DataCacheManager:
    """Sistema inteligente de cache para datos de acciones"""
    
    def __init__(self, cache_dir="cache"):
        self.cache_dir = Path(cache_dir)
        self.cache_dir.mkdir(exist_ok=True)
        
        self.data_file = self.cache_dir / "stock_data.parquet"
        self.metadata_file = self.cache_dir / "metadata.json"
        self.failed_file = self.cache_dir / "failed_symbols.txt"
    
    def get_cache_info(self):
        """Obtiene información del cache actual"""
        if not self.metadata_file.exists():
            return {
                'exists': False,
                'date': None,
                'symbols': [],
                'date_range': None
            }
        
        try:
            with open(self.metadata_file, 'r') as f:
                metadata = json.load(f)
            
            cache_date = datetime.fromisoformat(metadata['created_date'])
            age_days = (datetime.now() - cache_date).days
            
            return {
                'exists': True,
                'date': cache_date,
                'age_days': age_days,
                'symbols': metadata['symbols'],
                'date_range': metadata['date_range'],
                'total_symbols': len(metadata['symbols'])
            }
        except Exception as e:
            print(f"Error leyendo metadata: {e}")
            return {'exists': False, 'date': None, 'symbols': [], 'date_range': None}
    
    def should_use_cache(self, requested_symbols, max_age_days=7):
        """Determina si usar cache existente"""
        cache_info = self.get_cache_info()
        
        if not cache_info['exists']:
            return False, "No existe cache"
        
        if cache_info['age_days'] > max_age_days:
            return False, f"Cache muy antiguo ({cache_info['age_days']} días)"
        
        # Verificar si todos los símbolos están en cache
        cached_symbols = set(cache_info['symbols'])
        requested_set = set(requested_symbols)
        missing_symbols = requested_set - cached_symbols
        
        if missing_symbols:
            return False, f"Faltan símbolos: {list(missing_symbols)[:5]}..."
        
        return True, f"Cache válido ({cache_info['age_days']} días, {len(cached_symbols)} símbolos)"
    
    def load_cached_data(self, symbols=None):
        """Carga datos del cache"""
        try:
            if not self.data_file.exists():
                return None
            
            data = pd.read_parquet(self.data_file)
            
            if symbols:
                # Filtrar solo símbolos solicitados
                available_symbols = [col for col in data.columns if col in symbols]
                data = data[available_symbols]
            
            return data
        except Exception as e:
            print(f"Error cargando cache: {e}")
            return None
    
    def save_data_to_cache(self, data_dict, symbols, date_range):
        """Guarda datos en cache"""
        try:
            # Convertir dict a DataFrame
            df = pd.DataFrame(data_dict)
            df.to_parquet(self.data_file, compression='snappy')
            
            # Guardar metadata
            metadata = {
                'created_date': datetime.now().isoformat(),
                'symbols': symbols,
                'date_range': date_range,
                'total_symbols': len(symbols)
            }
            
            with open(self.metadata_file, 'w') as f:
                json.dump(metadata, f, indent=2)
            
            print(f"✅ Cache guardado: {len(symbols)} símbolos")
            return True
        except Exception as e:
            print(f"Error guardando cache: {e}")
            return False
    
    def get_cache_status_summary(self):
        """Resumen del estado del cache para mostrar al usuario"""
        cache_info = self.get_cache_info()
        
        if not cache_info['exists']:
            return "📂 Cache: No existe - se descargará todo"
        
        age = cache_info['age_days']
        count = cache_info['total_symbols']
        
        if age == 0:
            age_str = "hoy"
        elif age == 1:
            age_str = "ayer"
        else:
            age_str = f"{age} días"
        
        return f"📂 Cache: {count} símbolos ({age_str})"

class StockProfileClassifier:
    """Clasificador de perfiles de inversión para acciones"""
    
    @staticmethod
    def classify_historical_profile(metrics):
        """Clasifica perfil para análisis 5Y y 10Y"""
        cagr = metrics.get('cagr', 0)
        sharpe = metrics.get('sharpe', 0)
        volatility = metrics.get('volatility', 100)
        consistency = metrics.get('benchmark_win_rate', 0)
        
        # 🚀 GROWTH STAR: Alto crecimiento, acepta volatilidad
        if cagr >= 30 and sharpe >= 1.5 and consistency >= 0.8:
            return "🚀 GROWTH_STAR"
        
        # 💎 QUALITY CHAMPION: Balance perfecto
        elif cagr >= 25 and sharpe >= 1.8 and volatility <= 25 and consistency >= 0.8:
            return "💎 QUALITY_CHAMPION"
        
        # 🛡️ DEFENSIVE WINNER: Estabilidad + rendimiento decente
        elif cagr >= 18 and sharpe >= 1.6 and volatility <= 20 and consistency >= 0.75:
            return "🛡️ DEFENSIVE_WINNER"
        
        # ⚡ MOMENTUM BEAST: Crecimiento explosivo pero volátil
        elif cagr >= 35 and consistency >= 0.7:  # Permite más volatilidad
            return "⚡ MOMENTUM_BEAST"
        
        # 🎯 CONSISTENT PERFORMER: Gana siempre, rendimiento moderado
        elif cagr >= 20 and consistency >= 0.9 and sharpe >= 1.3:
            return "🎯 CONSISTENT_PERFORMER"
        
        # 💰 VALUE COMPOUNDER: Crecimiento sostenido, bajo riesgo
        elif cagr >= 22 and volatility <= 18 and sharpe >= 1.5:
            return "💰 VALUE_COMPOUNDER"
        
        # 🔥 TURNAROUND STORY: Mejorando fuertemente
        elif cagr >= 28 and sharpe >= 1.2:  # Menos estricto en otros criterios
            return "🔥 TURNAROUND_STORY"
        
        # 📈 SOLID PERFORMER: Buena pero sin destacar
        else:
            return "📈 SOLID_PERFORMER"
    
    @staticmethod
    def classify_ytd_profile(metrics):
        """Clasifica perfil para análisis YTD"""
        ytd_return = metrics.get('ytd_return', 0)
        volatility = metrics.get('volatility', 100)
        sharpe = metrics.get('sharpe', 0)
        beats_benchmarks = metrics.get('benchmark_wins', 0)
        
        # 🔥 YTD ROCKET: Rendimiento explosivo
        if ytd_return >= 50:
            return "🔥 YTD_ROCKET"
        
        # 🚀 STRONG MOMENTUM: Muy buen año
        elif ytd_return >= 30 and beats_benchmarks >= 2:
            return "🚀 STRONG_MOMENTUM"
        
        # 💎 BALANCED WINNER: Buen año equilibrado
        elif ytd_return >= 20 and volatility <= 30 and beats_benchmarks >= 2:
            return "💎 BALANCED_WINNER"
        
        # 🛡️ DEFENSIVE OUTPERFORMER: Gana sin mucho riesgo
        elif ytd_return >= 15 and volatility <= 20 and beats_benchmarks >= 1:
            return "🛡️ DEFENSIVE_OUTPERFORMER"
        
        # ⚡ VOLATILE WINNER: Gana mucho pero con riesgo
        elif ytd_return >= 25 and volatility >= 35:
            return "⚡ VOLATILE_WINNER"
        
        # 📊 BENCHMARK BEATER: Supera mercado consistentemente
        elif ytd_return >= 10 and beats_benchmarks >= 2:
            return "📊 BENCHMARK_BEATER"
        
        # 🎯 STEADY CLIMBER: Crecimiento constante
        elif ytd_return >= 15 and beats_benchmarks >= 1:
            return "🎯 STEADY_CLIMBER"
        
        # 📈 MARKET PERFORMER: Sigue al mercado
        else:
            return "📈 MARKET_PERFORMER"
    
    @staticmethod
    def calculate_balanced_score(metrics):
        """Calcula score balanceado 0-100 para ranking"""
        # Normalizar cada componente a 0-25
        
        # CAGR: 0-40% → 0-25 puntos
        cagr_score = min(metrics.get('cagr', 0) * 25 / 40, 25)
        
        # Sharpe: 0-2.5 → 0-25 puntos
        sharpe_score = min(metrics.get('sharpe', 0) * 25 / 2.5, 25)
        
        # Consistencia: 0-1 → 0-25 puntos
        consistency = metrics.get('benchmark_win_rate', 0)
        consistency_score = consistency * 25
        
        # Estabilidad: volatilidad 15-50% → 25-0 puntos
        volatility = metrics.get('volatility', 50)
        volatility_score = max(0, 25 - (volatility - 15) * 25 / 35)
        
        total_score = cagr_score + sharpe_score + consistency_score + volatility_score
        
        return {
            'total_score': round(total_score, 1),
            'breakdown': {
                'cagr': round(cagr_score, 1),
                'sharpe': round(sharpe_score, 1),
                'consistency': round(consistency_score, 1),
                'stability': round(volatility_score, 1)
            }
        }

class DataFetcher:
    """Módulo para obtener datos históricos de acciones y ETFs"""
    
    def __init__(self, start_date, end_date):
        self.start_date = start_date
        self.end_date = end_date
    
    def fetch_stock_data(self, symbol):
        """Obtiene datos históricos para un símbolo específico"""
        try:
            stock = yf.Ticker(symbol)
            data = stock.history(start=self.start_date, end=self.end_date)
            if data.empty:
                print(f"Sin datos para {symbol}")
                return None
            return data['Close']
        except Exception as e:
            print(f"Error obteniendo {symbol}: {e}")
            return None
    
    def fetch_multiple_stocks(self, symbols, batch_size=100):
        """Obtiene datos para múltiples símbolos de forma eficiente"""
        print(f"Descargando datos para {len(symbols)} símbolos desde {self.start_date} hasta {self.end_date}...")
        
        results = {}
        total_batches = (len(symbols) + batch_size - 1) // batch_size
        
        for i in range(0, len(symbols), batch_size):
            batch_symbols = symbols[i:i + batch_size]
            batch_num = i // batch_size + 1
            
            print(f"Procesando lote {batch_num}/{total_batches} ({len(batch_symbols)} símbolos)")
            
            try:
                # Filtrar símbolos válidos
                valid_symbols = [s for s in batch_symbols if isinstance(s, str) and len(s) > 0]
                
                if not valid_symbols:
                    continue
                
                symbols_str = " ".join(valid_symbols)
                data = yf.download(symbols_str, start=self.start_date, end=self.end_date, 
                                 group_by='ticker', progress=False, threads=True)
                
                for symbol in valid_symbols:
                    try:
                        if len(valid_symbols) == 1:
                            stock_data = data['Close'] if 'Close' in data.columns else data
                        else:
                            if symbol in data.columns.get_level_values(0):
                                stock_data = data[symbol]['Close']
                            else:
                                print(f"No se encontraron datos para {symbol}")
                                continue
                        
                        if not stock_data.empty and not stock_data.isna().all():
                            # Verificar que tengamos al menos 1 año de datos
                            if len(stock_data.dropna()) > 50:  # ~50 días de trading por año mínimo
                                results[symbol] = stock_data.dropna()
                            else:
                                print(f"Datos insuficientes para {symbol}")
                        else:
                            print(f"Datos vacíos para {symbol}")
                            
                    except Exception as e:
                        print(f"Error procesando {symbol}: {e}")
                        continue
                        
            except Exception as e:
                print(f"Error en lote {batch_num}: {e}")
                continue
        
        print(f"✅ Datos obtenidos exitosamente para {len(results)} de {len(symbols)} símbolos")
        return results

class OptimizedDataFetcher:
    """Descarga optimizada con cache inteligente"""
    
    def __init__(self):
        self.cache_manager = DataCacheManager()
        self.date_ranges = DateManager.get_dynamic_date_ranges()
    
    def fetch_stock_data_optimized(self, symbols, force_download=False):
        """Descarga optimizada con sistema de cache"""
        
        print(f"\n🚀 DESCARGA OPTIMIZADA DE DATOS")
        print(f"📊 Símbolos solicitados: {len(symbols)}")
        print(f"📅 Rango: {self.date_ranges['full_download'][0]} → {self.date_ranges['full_download'][1]}")
        
        # Verificar cache
        cache_status = self.cache_manager.get_cache_status_summary()
        print(cache_status)
        
        should_use, reason = self.cache_manager.should_use_cache(symbols)
        
        if should_use and not force_download:
            print(f"✅ Usando cache: {reason}")
            cached_data = self.cache_manager.load_cached_data(symbols)
            if cached_data is not None:
                print(f"📈 Datos cargados del cache: {len(cached_data.columns)} símbolos")
                return self._convert_df_to_dict(cached_data)
        
        print(f"📥 Descargando datos: {reason}")
        return self._download_and_cache_data(symbols)
    
    def _download_and_cache_data(self, symbols):
        """Descarga datos y los guarda en cache"""
        start_date, end_date = self.date_ranges['full_download']
        
        print(f"\n💾 Descargando {len(symbols)} símbolos...")
        print(f"📅 Período: {start_date} → {end_date}")
        
        # Usar el fetcher original pero optimizado
        original_fetcher = DataFetcher(start_date, end_date)
        stock_data = original_fetcher.fetch_multiple_stocks(symbols)
        
        if stock_data:
            # Guardar en cache
            date_range = f"{start_date}_to_{end_date}"
            symbols_list = list(stock_data.keys())
            
            success = self.cache_manager.save_data_to_cache(
                stock_data, symbols_list, date_range
            )
            
            if success:
                print(f"💾 Cache actualizado exitosamente")
        
        return stock_data
    
    def _convert_df_to_dict(self, df):
        """Convierte DataFrame del cache a diccionario compatible"""
        result = {}
        for column in df.columns:
            series_data = df[column].dropna()
            if not series_data.empty:
                result[column] = series_data
        return result

class PerformanceCalculator:
    """Módulo para calcular métricas de rendimiento"""
    
    @staticmethod
    def calculate_annual_returns(price_data, start_year, end_year):
        """Calcula rendimientos anuales a partir de datos de precios"""
        if price_data is None or price_data.empty:
            return {}
        
        annual_returns = {}
        price_data = price_data.dropna()
        
        for year in range(start_year, end_year + 1):
            try:
                # Obtener datos del año
                year_data = price_data[price_data.index.year == year]
                
                if len(year_data) < 10:  # Mínimo 10 observaciones por año
                    continue
                
                # Precio al inicio y final del año
                start_price = year_data.iloc[0]
                end_price = year_data.iloc[-1]
                
                if start_price > 0:
                    annual_return = (end_price / start_price - 1) * 100
                    annual_returns[year] = annual_return
                    
            except Exception as e:
                print(f"Error calculando rendimiento para {year}: {e}")
                continue
        
        return annual_returns
    
    @staticmethod
    def calculate_total_return(price_data):
        """Calcula el rendimiento total del período"""
        if price_data is None or len(price_data) < 2:
            return 0
        
        try:
            start_price = price_data.iloc[0]
            end_price = price_data.iloc[-1]
            
            if start_price > 0:
                return (end_price / start_price - 1) * 100
            return 0
        except:
            return 0
    
    @staticmethod
    def calculate_cagr(price_data, years):
        """Calcula la tasa de crecimiento anual compuesta (CAGR)"""
        if price_data is None or len(price_data) < 2 or years <= 0:
            return 0
        
        try:
            start_price = price_data.iloc[0]
            end_price = price_data.iloc[-1]
            
            if start_price > 0:
                return (((end_price / start_price) ** (1/years)) - 1) * 100
            return 0
        except:
            return 0
    
    @staticmethod
    def calculate_sharpe_ratio(returns, risk_free_rate=0.02):
        """Calcula el ratio de Sharpe"""
        if not returns or len(returns) < 2:
            return 0
        
        try:
            returns_array = np.array(list(returns.values()))
            excess_returns = returns_array - risk_free_rate * 100
            
            if np.std(returns_array) == 0:
                return 0
            
            return np.mean(excess_returns) / np.std(returns_array)
        except:
            return 0
    
    @staticmethod
    def calculate_max_drawdown(price_data):
        """Calcula el drawdown máximo"""
        if price_data is None or len(price_data) < 2:
            return 0
        
        try:
            cumulative = (1 + price_data.pct_change().fillna(0)).cumprod()
            running_max = cumulative.expanding().max()
            drawdown = (cumulative - running_max) / running_max
            
            return abs(drawdown.min()) * 100
        except:
            return 0
    
    @staticmethod
    def calculate_volatility(price_data):
        """Calcula la volatilidad anualizada"""
        if price_data is None or len(price_data) < 2:
            return 0
        
        try:
            returns = price_data.pct_change().dropna()
            if len(returns) == 0:
                return 0
            return returns.std() * np.sqrt(252) * 100  # Anualizada
        except:
            return 0

class BenchmarkComparator:
    """Módulo para comparar rendimiento contra benchmarks"""
    
    def __init__(self, benchmark_returns):
        self.benchmark_returns = benchmark_returns
    
    def compare_performance(self, stock_returns, benchmark_name):
        """Compara rendimiento de una acción contra un benchmark"""
        if not stock_returns or benchmark_name not in self.benchmark_returns:
            return {}
        
        comparison = {}
        benchmark_data = self.benchmark_returns[benchmark_name]
        
        for year in stock_returns:
            if year in benchmark_data:
                stock_ret = stock_returns[year]
                bench_ret = benchmark_data[year]
                
                comparison[year] = {
                    'stock_return': stock_ret,
                    'benchmark_return': bench_ret,
                    'outperformance': stock_ret - bench_ret,
                    'beats_benchmark': stock_ret > bench_ret
                }
        
        return comparison
    
    def count_outperformance_years(self, stock_returns, benchmark_name):
        """Cuenta años de outperformance contra un benchmark"""
        comparison = self.compare_performance(stock_returns, benchmark_name)
        
        if not comparison:
            return 0, 0, 0
        
        total_years = len(comparison)
        outperform_years = sum(1 for data in comparison.values() if data['beats_benchmark'])
        avg_outperformance = np.mean([data['outperformance'] for data in comparison.values()])
        
        return outperform_years, total_years, avg_outperformance

class OutperformanceFilter:
    """Módulo para filtrar acciones con outperformance sostenida - MEJORADO"""
    
    def __init__(self, comparator, min_years_percentage=0.7):  # Cambiado a 0.7
        self.comparator = comparator
        self.min_years_percentage = min_years_percentage
    
    def filter_sustained_outperformers(self, stock_returns_dict, benchmarks, total_years):
        """Filtra acciones que superan MAYORÍA de benchmarks en 70% del tiempo"""
        min_years = max(1, int(total_years * self.min_years_percentage))
        sustained_outperformers = []
        
        print(f"Filtros MEJORADOS: Mínimo {min_years} de {total_years} años + superar mayoría de benchmarks")
        
        for symbol, returns in stock_returns_dict.items():
            if not returns:
                continue
            
            # Contar outperformance por benchmark
            benchmark_wins = 0
            performance_summary = {'symbol': symbol}
            
            for benchmark in benchmarks:
                outperform_years, stock_total_years, avg_outperformance = \
                    self.comparator.count_outperformance_years(returns, benchmark)
                
                performance_summary[f'beats_{benchmark}_years'] = outperform_years
                performance_summary[f'total_years'] = stock_total_years
                performance_summary[f'{benchmark}_avg_outperformance'] = avg_outperformance
                
                # Cuenta si supera este benchmark el 70% del tiempo
                if outperform_years >= min_years:
                    benchmark_wins += 1
            
            # NUEVO CRITERIO: Debe superar al menos 2 de 3 benchmarks
            beats_majority = benchmark_wins >= 2
            
            # Verificar Sharpe superior (si disponible)
            # Esto se calculará en el análisis principal
            
            if beats_majority:
                performance_summary['benchmarks_beaten'] = benchmark_wins
                sustained_outperformers.append(performance_summary)
        
        print(f"✅ Outperformers encontrados: {len(sustained_outperformers)} (criterio: mayoría de benchmarks)")
        return sustained_outperformers

class EnhancedStockAnalyzer:
    """Analizador mejorado con descarga única y análisis múltiple"""
    
    def __init__(self, excel_file='Market_Cap_Ranked.xlsx'):
        self.excel_file = excel_file
        self.benchmarks = ['SPY', 'QQQ', 'SPYG']
        self.stock_symbols = self._load_stock_symbols()
        self.fetcher = OptimizedDataFetcher()
        self.analysis_periods = DateManager.get_analysis_periods()
        
        # Cache para datos descargados
        self._cached_stock_data = None
        self._cached_benchmark_data = None
    
    def _load_stock_symbols(self):
        """Carga símbolos desde Excel (reutiliza lógica original)"""
        try:
            print(f"📊 Cargando símbolos desde {self.excel_file}...")
            df = pd.read_excel(self.excel_file)
            
            if 'Ticker' not in df.columns:
                raise ValueError("La columna 'Ticker' no se encontró en el Excel")
            
            if 'Market Cap Rank' not in df.columns:
                print("⚠️ Advertencia: 'Market Cap Rank' no encontrado, usando orden del archivo")
                df['Market Cap Rank'] = range(1, len(df) + 1)
            
            df_sorted = df.sort_values('Market Cap Rank').head(5769)
            symbols = df_sorted['Ticker'].dropna().unique().tolist()
            symbols = [str(symbol).strip().upper() for symbol in symbols if pd.notna(symbol) and str(symbol).strip()]
            
            print(f"✅ {len(symbols)} símbolos cargados exitosamente")
            return symbols
            
        except Exception as e:
            print(f"❌ Error cargando Excel: {e}")
            print("Usando símbolos de respaldo...")
            return [
                'AAPL', 'MSFT', 'NVDA', 'AMZN', 'META', 'GOOGL', 'AVGO', 'TSLA', 'GOOG',
                'LLY', 'JPM', 'V', 'NFLX', 'XOM', 'MA', 'COST', 'WMT', 'PG', 'UNH',
                'JNJ', 'HD', 'ABBV', 'KO', 'PM', 'BAC', 'CRM', 'WFC', 'CSCO',
                'MCD', 'ORCL', 'CVX', 'ABT', 'IBM', 'GE', 'LIN', 'MRK', 'T'
            ]
    
    def download_all_data(self, force_refresh=False):
        """Descarga TODOS los datos de una vez (stock + benchmarks)"""
        
        print(f"\n{'='*60}")
        print("📥 DESCARGA ÚNICA DE TODOS LOS DATOS")
        print(f"{'='*60}")
        
        # 1. Descargar datos de acciones
        print(f"\n1️⃣ Descargando datos de acciones...")
        self._cached_stock_data = self.fetcher.fetch_stock_data_optimized(
            self.stock_symbols, force_download=force_refresh
        )
        
        # 2. Descargar benchmarks
        print(f"\n2️⃣ Descargando benchmarks...")
        self._cached_benchmark_data = self.fetcher.fetch_stock_data_optimized(
            self.benchmarks, force_download=force_refresh
        )
        
        # Resumen de descarga
        stock_count = len(self._cached_stock_data) if self._cached_stock_data else 0
        benchmark_count = len(self._cached_benchmark_data) if self._cached_benchmark_data else 0
        
        print(f"\n✅ DESCARGA COMPLETADA:")
        print(f"   📈 Acciones: {stock_count} símbolos")
        print(f"   📊 Benchmarks: {benchmark_count} símbolos")
        print(f"   💾 Datos en memoria: ✓")
        
        return stock_count > 0 and benchmark_count > 0
    
    def run_analysis_for_period(self, period_key, min_outperform_percentage=0.7):
        """Ejecuta análisis para un período específico usando datos cached"""
        
        if not self._cached_stock_data or not self._cached_benchmark_data:
            raise ValueError("Datos no descargados. Ejecuta download_all_data() primero.")
        
        period_info = self.analysis_periods[period_key]
        period_name = period_info['name']
        
        print(f"\n{'='*20} ANÁLISIS {period_name.upper()} {'='*20}")
        
        calculator = PerformanceCalculator()
        
        # 1. Procesar benchmarks para este período
        benchmark_returns = {}
        for benchmark, data in self._cached_benchmark_data.items():
            if data is None or data.empty:
                continue
            
            # Filtrar datos por período
            period_data = self._filter_data_by_period(data, period_key)
            if period_data is None:
                continue
            
            # Calcular retornos según el período
            if period_key == 'ytd':
                if len(period_data) >= 2:
                    start_price = period_data.iloc[0]
                    end_price = period_data.iloc[-1]
                    if start_price > 0:
                        ytd_return = (end_price / start_price - 1) * 100
                        current_year = datetime.now().year
                        benchmark_returns[benchmark] = {current_year: ytd_return}
            else:
                # Análisis histórico
                years_count = period_info['years_count']
                current_year = datetime.now().year
                start_year = current_year - years_count
                end_year = current_year - 1
                
                returns = calculator.calculate_annual_returns(period_data, start_year, end_year)
                if returns:
                    benchmark_returns[benchmark] = returns
        
        print(f"📊 Benchmarks procesados: {len(benchmark_returns)}")
        for bench, returns in benchmark_returns.items():
            if returns:
                avg_ret = np.mean(list(returns.values()))
                print(f"   {bench}: {avg_ret:.2f}% promedio")
        
        # 2. Procesar acciones para este período
        stock_returns = {}
        results = []
        
        print(f"🧮 Procesando {len(self._cached_stock_data)} acciones...")
        
        for symbol, data in self._cached_stock_data.items():
            if data is None or data.empty:
                continue
            
            # Filtrar datos por período
            period_data = self._filter_data_by_period(data, period_key)
            if period_data is None or len(period_data) < 10:
                continue
            
            # Calcular retornos según período
            if period_key == 'ytd':
                if len(period_data) >= 2:
                    start_price = period_data.iloc[0]
                    end_price = period_data.iloc[-1]
                    if start_price > 0:
                        ytd_return = (end_price / start_price - 1) * 100
                        current_year = datetime.now().year
                        stock_returns[symbol] = {current_year: ytd_return}
            else:
                # Análisis histórico
                years_count = period_info['years_count']
                current_year = datetime.now().year
                start_year = current_year - years_count
                end_year = current_year - 1
                
                returns = calculator.calculate_annual_returns(period_data, start_year, end_year)
                min_years = max(1, int(years_count * 0.6))
                if returns and len(returns) >= min_years:
                    stock_returns[symbol] = returns
        
        print(f"📈 Acciones procesadas: {len(stock_returns)}")
        
        # 3. Análisis de outperformance
        if benchmark_returns:
            comparator = BenchmarkComparator(benchmark_returns)
            
            if period_key == 'ytd':
                filter_obj = OutperformanceFilter(comparator, min_years_percentage=1.0)
                total_periods = 1
            else:
                filter_obj = OutperformanceFilter(comparator, min_outperform_percentage)
                total_periods = period_info['years_count']
            
            outperformers = filter_obj.filter_sustained_outperformers(
                stock_returns, self.benchmarks, total_periods
            )
        else:
            outperformers = []
        
        # 4. Calcular métricas adicionales y clasificación
        classifier = StockProfileClassifier()
        
        for symbol in stock_returns.keys():
            if symbol in self._cached_stock_data:
                period_data = self._filter_data_by_period(self._cached_stock_data[symbol], period_key)
                returns = stock_returns[symbol]
                
                # Calcular métricas según el período
                if period_key == 'ytd':
                    main_return = list(returns.values())[0] if returns else 0
                    total_return = main_return
                    cagr = main_return
                    period_suffix = "YTD"
                else:
                    main_return = np.mean(list(returns.values())) if returns else 0
                    total_return = calculator.calculate_total_return(period_data)
                    cagr = calculator.calculate_cagr(period_data, period_info['years_count'])
                    period_suffix = f"{period_info['years_count']}Y"
                
                # Métricas de riesgo
                sharpe = calculator.calculate_sharpe_ratio(returns) if returns else 0
                max_dd = calculator.calculate_max_drawdown(period_data)
                volatility = calculator.calculate_volatility(period_data)
                
                # Outperformance contra benchmarks
                if benchmark_returns:
                    spy_years, spy_total, spy_outperf = comparator.count_outperformance_years(returns, 'SPY')
                    qqq_years, qqq_total, qqq_outperf = comparator.count_outperformance_years(returns, 'QQQ')
                    spyg_years, spyg_total, spyg_outperf = comparator.count_outperformance_years(returns, 'SPYG')
                else:
                    spy_years = qqq_years = spyg_years = 0
                    spy_outperf = qqq_outperf = spyg_outperf = 0
                
                # NUEVA FUNCIONALIDAD: Clasificación y Score
                if period_key == 'ytd':
                    # Clasificación YTD
                    ytd_metrics = {
                        'ytd_return': main_return,
                        'volatility': volatility,
                        'sharpe': sharpe,
                        'benchmark_wins': sum([spy_years > 0, qqq_years > 0, spyg_years > 0])
                    }
                    investment_profile = classifier.classify_ytd_profile(ytd_metrics)
                    quality_score = 0  # No aplicamos score histórico al YTD
                else:
                    # Clasificación histórica (5Y/10Y)
                    total_years = period_info['years_count']
                    benchmark_win_rate = (spy_years + qqq_years + spyg_years) / (3 * total_years)
                    
                    historical_metrics = {
                        'cagr': cagr,
                        'sharpe': sharpe,
                        'volatility': volatility,
                        'benchmark_win_rate': benchmark_win_rate
                    }
                    investment_profile = classifier.classify_historical_profile(historical_metrics)
                    
                    # Calcular score balanceado
                    score_data = classifier.calculate_balanced_score(historical_metrics)
                    quality_score = score_data['total_score']
                
                result = {
                    'Symbol': symbol,
                    f'Return_{period_suffix}_%': round(main_return, 2),
                    f'Total_Return_{period_suffix}_%': round(total_return, 2),
                    f'CAGR_{period_suffix}_%': round(cagr, 2),
                    'Beats_SPY': spy_years,
                    'Beats_QQQ': qqq_years,
                    'Beats_SPYG': spyg_years,
                    'SPY_Outperformance_%': round(spy_outperf, 2),
                    'QQQ_Outperformance_%': round(qqq_outperf, 2),
                    'SPYG_Outperformance_%': round(spyg_outperf, 2),
                    'Sharpe_Ratio': round(sharpe, 2),
                    'Max_Drawdown_%': round(max_dd, 2),
                    'Volatility_%': round(volatility, 2),
                    'Investment_Profile': investment_profile,
                    'Quality_Score': round(quality_score, 1) if period_key != 'ytd' else 'N/A',
                    'Data_Periods': len(returns) if returns else 0
                }
                results.append(result)
        
        print(f"✅ {period_name}: {len(results)} resultados, {len(outperformers)} outperformers")
        
        return results, benchmark_returns, outperformers
    
    def _filter_data_by_period(self, data, period_key):
        """Filtra datos según el período solicitado"""
        if data is None or data.empty:
            return None
        
        try:
            period_info = self.analysis_periods[period_key]
            start_date = pd.to_datetime(period_info['start_date'])
            end_date = pd.to_datetime(period_info['end_date'])
            
            # Filtrar por fechas
            mask = (data.index >= start_date) & (data.index <= end_date)
            filtered_data = data[mask]
            
            return filtered_data if not filtered_data.empty else None
        except Exception as e:
            print(f"Error filtrando datos para {period_key}: {e}")
            return None
    
    def run_complete_analysis(self, force_refresh=False):
        """Ejecuta análisis completo: descarga única + análisis múltiple"""
        
        print(f"\n🚀 ANÁLISIS COMPLETO - ARQUITECTURA OPTIMIZADA")
        print(f"{'='*70}")
        
        # 1. Descarga única de todos los datos
        download_success = self.download_all_data(force_refresh)
        if not download_success:
            print("❌ Error en descarga de datos")
            return None
        
        # 2. Análisis múltiple usando los mismos datos
        all_results = {}
        all_benchmarks = {}
        all_outperformers = {}
        
        # Orden de prioridad: YTD, 5Y, 10Y
        analysis_order = ['ytd', '5y', '10y']
        
        for period_key in analysis_order:
            try:
                results, benchmarks, outperformers = self.run_analysis_for_period(period_key)
                
                all_results[period_key] = results
                all_benchmarks[period_key] = benchmarks
                all_outperformers[period_key] = outperformers
                
            except Exception as e:
                print(f"❌ Error en análisis {period_key}: {e}")
                all_results[period_key] = []
                all_benchmarks[period_key] = {}
                all_outperformers[period_key] = []
        
        # 3. Generar Excel
        print(f"\n📊 Generando reporte Excel...")
        filename = EnhancedExcelExporter.create_comprehensive_excel(
            all_results, all_benchmarks, all_outperformers
        )
        
        # 4. Generar recomendaciones DCA
        self._generate_dca_recommendations(all_results)
        
        print(f"\n✅ ANÁLISIS COMPLETADO")
        print(f"📄 Archivo generado: {filename}")
        
        return filename, all_results, all_benchmarks, all_outperformers

class EnhancedExcelExporter:
    """Exportador mejorado para múltiples análisis"""
    
    @staticmethod
    def create_comprehensive_excel(all_results, all_benchmarks, all_outperformers, 
                                 filename="Stock_Analysis_Enhanced.xlsx"):
        """Crea Excel completo con todos los análisis"""
        
        try:
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                
                # Prioridad 1: YTD Analysis
                if 'ytd' in all_results and all_results['ytd']:
                    df_ytd = pd.DataFrame(all_results['ytd'])
                    df_ytd = df_ytd.sort_values('Return_YTD_%', ascending=False)
                    df_ytd.to_excel(writer, sheet_name='YTD_2025_Analysis', index=False)
                    
                    ws_ytd = writer.sheets['YTD_2025_Analysis']
                    EnhancedExcelExporter._format_worksheet(ws_ytd, "🔥 Análisis YTD 2025")
                
                # Análisis 5 años
                if '5y' in all_results and all_results['5y']:
                    df_5y = pd.DataFrame(all_results['5y'])
                    # NUEVO: Ordenar por Quality Score en lugar de solo CAGR
                    df_5y = df_5y.sort_values(['Quality_Score', 'CAGR_5Y_%'], ascending=[False, False])
                    df_5y.to_excel(writer, sheet_name='Analysis_5_Years', index=False)
                    
                    ws_5y = writer.sheets['Analysis_5_Years']
                    EnhancedExcelExporter._format_worksheet(ws_5y, "📊 Análisis 5 Años (Ordenado por Quality Score)")
                
                # Análisis 10 años
                if '10y' in all_results and all_results['10y']:
                    df_10y = pd.DataFrame(all_results['10y'])
                    # NUEVO: Ordenar por Quality Score en lugar de solo CAGR
                    df_10y = df_10y.sort_values(['Quality_Score', 'CAGR_10Y_%'], ascending=[False, False])
                    df_10y.to_excel(writer, sheet_name='Analysis_10_Years', index=False)
                    
                    ws_10y = writer.sheets['Analysis_10_Years']
                    EnhancedExcelExporter._format_worksheet(ws_10y, "📈 Análisis 10 Años (Ordenado por Quality Score)")
                
                # NUEVA HOJA: Top Investment Profiles
                EnhancedExcelExporter._create_profiles_summary_sheet(writer, all_results)
                
                # Outperformers sheets - MEJORADOS con perfiles
                for period_key, period_name in [('ytd', 'YTD'), ('5y', '5Y'), ('10y', '10Y')]:
                    if (period_key in all_outperformers and all_outperformers[period_key] and 
                        period_key in all_results and all_results[period_key]):
                        
                        outperf_symbols = [item['symbol'] for item in all_outperformers[period_key]]
                        outperf_data = [r for r in all_results[period_key] if r['Symbol'] in outperf_symbols]
                        
                        if outperf_data:
                            df_outperf = pd.DataFrame(outperf_data)
                            
                            # Ordenamiento inteligente
                            if period_key == 'ytd':
                                df_outperf = df_outperf.sort_values('Return_YTD_%', ascending=False)
                            else:
                                # Ordenar por Quality Score para análisis históricos
                                df_outperf = df_outperf.sort_values(['Quality_Score', f'CAGR_{period_name}_%'], 
                                                                  ascending=[False, False])
                            
                            sheet_name = f'Outperformers_{period_name}'
                            df_outperf.to_excel(writer, sheet_name=sheet_name, index=False)
                            
                            ws_out = writer.sheets[sheet_name]
                            title = f"🚀 Outperformers {period_name} (Mayoría de Benchmarks + 70%)"
                            EnhancedExcelExporter._format_worksheet(ws_out, title, highlight=True)
                
                # Summary sheet
                EnhancedExcelExporter._create_summary_sheet(writer, all_results, all_benchmarks)
            
            print(f"✅ Excel creado: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ Error creando Excel: {e}")
            return None
    
    @staticmethod
    def _create_profiles_summary_sheet(writer, all_results):
        """Crea hoja resumen de perfiles de inversión"""
        
        profiles_data = []
        
        # Procesar análisis históricos (5Y y 10Y)
        for period_key, period_name in [('5y', '5Y'), ('10y', '10Y')]:
            if period_key in all_results and all_results[period_key]:
                # Agrupar por perfil de inversión
                profile_groups = {}
                
                for result in all_results[period_key]:
                    profile = result.get('Investment_Profile', 'Unknown')
                    if profile not in profile_groups:
                        profile_groups[profile] = []
                    profile_groups[profile].append(result)
                
                # Crear resumen por perfil
                for profile, stocks in profile_groups.items():
                    if len(stocks) >= 1:  # Al menos 1 acción en el perfil
                        # Calcular estadísticas del grupo
                        avg_cagr = np.mean([s.get(f'CAGR_{period_name}_%', 0) for s in stocks])
                        avg_sharpe = np.mean([s.get('Sharpe_Ratio', 0) for s in stocks])
                        avg_volatility = np.mean([s.get('Volatility_%', 0) for s in stocks])
                        avg_score = np.mean([s.get('Quality_Score', 0) for s in stocks if s.get('Quality_Score') != 'N/A'])
                        
                        # Top 3 acciones del perfil
                        top_stocks = sorted(stocks, key=lambda x: x.get('Quality_Score', 0), reverse=True)[:3]
                        top_symbols = [s['Symbol'] for s in top_stocks]
                        
                        profiles_data.append({
                            'Period': period_name,
                            'Investment_Profile': profile,
                            'Stock_Count': len(stocks),
                            'Avg_CAGR_%': round(avg_cagr, 2),
                            'Avg_Sharpe': round(avg_sharpe, 2),
                            'Avg_Volatility_%': round(avg_volatility, 2),
                            'Avg_Quality_Score': round(avg_score, 1),
                            'Top_3_Stocks': ', '.join(top_symbols)
                        })
        
        # Procesar YTD
        if 'ytd' in all_results and all_results['ytd']:
            ytd_profile_groups = {}
            
            for result in all_results['ytd']:
                profile = result.get('Investment_Profile', 'Unknown')
                if profile not in ytd_profile_groups:
                    ytd_profile_groups[profile] = []
                ytd_profile_groups[profile].append(result)
            
            for profile, stocks in ytd_profile_groups.items():
                if len(stocks) >= 1:
                    avg_ytd = np.mean([s.get('Return_YTD_%', 0) for s in stocks])
                    avg_volatility = np.mean([s.get('Volatility_%', 0) for s in stocks])
                    
                    top_stocks = sorted(stocks, key=lambda x: x.get('Return_YTD_%', 0), reverse=True)[:3]
                    top_symbols = [s['Symbol'] for s in top_stocks]
                    
                    profiles_data.append({
                        'Period': 'YTD',
                        'Investment_Profile': profile,
                        'Stock_Count': len(stocks),
                        'Avg_CAGR_%': round(avg_ytd, 2),  # YTD return en lugar de CAGR
                        'Avg_Sharpe': 'N/A',
                        'Avg_Volatility_%': round(avg_volatility, 2),
                        'Avg_Quality_Score': 'N/A',
                        'Top_3_Stocks': ', '.join(top_symbols)
                    })
        
        if profiles_data:
            df_profiles = pd.DataFrame(profiles_data)
            # Ordenar por período y luego por score/rendimiento
            df_profiles = df_profiles.sort_values(['Period', 'Avg_Quality_Score'], 
                                                ascending=[True, False], na_position='last')
            
            df_profiles.to_excel(writer, sheet_name='Investment_Profiles', index=False)
            
            ws_profiles = writer.sheets['Investment_Profiles']
            EnhancedExcelExporter._format_worksheet(ws_profiles, "🏷️ Perfiles de Inversión por Categoría")
    
    @staticmethod
    def _create_summary_sheet(writer, all_results, all_benchmarks):
        """Crea hoja de resumen con benchmarks y top performers"""
        
        summary_data = []
        
        # Benchmarks
        for period_key, period_name in [('ytd', 'YTD'), ('5y', '5Y'), ('10y', '10Y')]:
            if period_key in all_benchmarks:
                for benchmark, returns in all_benchmarks[period_key].items():
                    if returns:
                        avg_return = np.mean(list(returns.values()))
                        summary_data.append({
                            'Category': 'Benchmark',
                            'Symbol': benchmark,
                            'Period': period_name,
                            'Return_%': round(avg_return, 2),
                            'Type': 'ETF'
                        })
        
        # Top performers
        for period_key, period_name in [('ytd', 'YTD'), ('5y', '5Y'), ('10y', '10Y')]:
            if period_key in all_results and all_results[period_key]:
                sort_col = 'Return_YTD_%' if period_key == 'ytd' else f'CAGR_{period_name}_%'
                top_stocks = sorted(all_results[period_key], 
                                  key=lambda x: x.get(sort_col, 0), reverse=True)[:3]
                
                for i, stock in enumerate(top_stocks, 1):
                    summary_data.append({
                        'Category': f'Top_{i}',
                        'Symbol': stock['Symbol'],
                        'Period': period_name,
                        'Return_%': stock.get(sort_col, 0),
                        'Type': 'Stock'
                    })
        
        if summary_data:
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, sheet_name='Summary_All_Periods', index=False)
            
            ws_summary = writer.sheets['Summary_All_Periods']
            EnhancedExcelExporter._format_worksheet(ws_summary, "📊 Resumen: Benchmarks + Top Performers")
    
    @staticmethod
    def _format_worksheet(worksheet, title, highlight=False):
        """Aplica formato profesional a las hojas - MEJORADO"""
        
        # Insertar título
        worksheet.insert_rows(1)
        worksheet['A1'] = title
        worksheet['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        worksheet['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Formatear encabezados (ahora en fila 2)
        header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_font = Font(bold=True)
        
        for cell in worksheet[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # NUEVO: Formateo condicional para perfiles de inversión
        if 'Investment_Profile' in [cell.value for cell in worksheet[2]]:
            EnhancedExcelExporter._apply_profile_conditional_formatting(worksheet)
        
        # Destacar outperformers
        if highlight:
            highlight_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                for cell in row:
                    cell.fill = highlight_fill
        
        # Ajustar anchos de columna
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 25)  # Aumentado para perfiles
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _apply_profile_conditional_formatting(worksheet):
        """Aplica colores específicos a cada perfil de inversión"""
        
        # Mapeo de perfiles a colores
        profile_colors = {
            '🚀 GROWTH_STAR': 'FF6B6B',      # Rojo brillante
            '💎 QUALITY_CHAMPION': '4ECDC4',  # Verde agua
            '🛡️ DEFENSIVE_WINNER': '95E1D3',  # Verde claro
            '⚡ MOMENTUM_BEAST': 'FFE66D',     # Amarillo
            '🎯 CONSISTENT_PERFORMER': 'A8E6CF', # Verde menta
            '💰 VALUE_COMPOUNDER': 'DDA0DD',   # Púrpura claro
            '🔥 TURNAROUND_STORY': 'FFA07A',   # Salmón
            '📈 SOLID_PERFORMER': 'E6E6FA',    # Lavanda
            
            # Perfiles YTD
            '🔥 YTD_ROCKET': 'FF4500',         # Rojo fuego
            '🚀 STRONG_MOMENTUM': 'FF6347',    # Tomate
            '💎 BALANCED_WINNER': '20B2AA',    # Verde mar
            '🛡️ DEFENSIVE_OUTPERFORMER': '98FB98', # Verde pálido
            '⚡ VOLATILE_WINNER': 'FFD700',     # Dorado
            '📊 BENCHMARK_BEATER': '87CEEB',    # Azul cielo
            '🎯 STEADY_CLIMBER': '90EE90',      # Verde claro
            '📈 MARKET_PERFORMER': 'F0F8FF'     # Azul Alice
        }
        
        # Encontrar columna de Investment_Profile
        profile_column = None
        for cell in worksheet[2]:
            if cell.value == 'Investment_Profile':
                profile_column = cell.column
                break
        
        if profile_column:
            # Aplicar colores por fila
            for row in worksheet.iter_rows(min_row=3, max_row=worksheet.max_row):
                profile_cell = row[profile_column - 1]  # Ajustar índice
                profile_value = profile_cell.value
                
                if profile_value in profile_colors:
                    color = profile_colors[profile_value]
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    profile_cell.fill = fill

def _generate_dca_recommendations(analyzer, all_results):
    """Genera recomendaciones DCA mejoradas con clasificación de perfiles"""
    
    print(f"\n{'='*60}")
    print("💰 RECOMENDACIONES ESTRATEGIA DCA - CON PERFILES")
    print(f"{'='*60}")
    
    if not all_results.get('ytd') or not all_results.get('5y') or not all_results.get('10y'):
        print("⚠️ Datos insuficientes para recomendaciones DCA")
        return
    
    # Combinar análisis de todos los períodos
    dca_candidates = []
    
    # Crear mapeo de símbolos
    ytd_data = {r['Symbol']: r for r in all_results['ytd']}
    y5_data = {r['Symbol']: r for r in all_results['5y']}
    y10_data = {r['Symbol']: r for r in all_results['10y']}
    
    # Encontrar símbolos con datos completos
    common_symbols = set(ytd_data.keys()) & set(y5_data.keys()) & set(y10_data.keys())
    
    print(f"📊 Analizando {len(common_symbols)} acciones con datos completos...")
    
    # Agrupar por perfil de inversión
    profile_groups = {
        'growth': [],
        'quality': [],
        'defensive': [],
        'momentum': [],
        'consistent': [],
        'value': []
    }
    
    for symbol in common_symbols:
        ytd = ytd_data[symbol]
        y5 = y5_data[symbol]
        y10 = y10_data[symbol]
        
        # Obtener perfil 10Y (más confiable para DCA)
        profile_10y = y10.get('Investment_Profile', '')
        quality_score = y10.get('Quality_Score', 0)
        
        # Criterios DCA mínimos
        cagr_10y = y10.get('CAGR_10Y_%', 0)
        sharpe_10y = y10.get('Sharpe_Ratio', 0)
        volatility_10y = y10.get('Volatility_%', 100)
        
        # Filtros básicos DCA
        if (cagr_10y > 15 and sharpe_10y > 0.8 and volatility_10y < 50 and quality_score > 40):
            
            candidate_data = {
                'symbol': symbol,
                'profile_10y': profile_10y,
                'quality_score': quality_score,
                'cagr_10y': cagr_10y,
                'cagr_5y': y5.get('CAGR_5Y_%', 0),
                'ytd_return': ytd.get('Return_YTD_%', 0),
                'sharpe_10y': sharpe_10y,
                'volatility_10y': volatility_10y,
                'max_dd_10y': y10.get('Max_Drawdown_%', 100),
                'beats_spy_10y': y10.get('Beats_SPY', 0)
            }
            
            # Clasificar por tipo de perfil
            if 'GROWTH' in profile_10y or 'MOMENTUM' in profile_10y:
                profile_groups['growth'].append(candidate_data)
            elif 'QUALITY' in profile_10y or 'CHAMPION' in profile_10y:
                profile_groups['quality'].append(candidate_data)
            elif 'DEFENSIVE' in profile_10y:
                profile_groups['defensive'].append(candidate_data)
            elif 'CONSISTENT' in profile_10y:
                profile_groups['consistent'].append(candidate_data)
            elif 'VALUE' in profile_10y:
                profile_groups['value'].append(candidate_data)
            else:
                profile_groups['quality'].append(candidate_data)  # Default a quality
    
    # Ordenar cada grupo por quality score
    for profile_type in profile_groups:
        profile_groups[profile_type].sort(key=lambda x: x['quality_score'], reverse=True)
    
    print(f"\n🏷️ DISTRIBUCIÓN POR PERFILES:")
    for profile_type, candidates in profile_groups.items():
        count = len(candidates)
        if count > 0:
            avg_score = np.mean([c['quality_score'] for c in candidates])
            print(f"   {profile_type.upper()}: {count} candidatos (Score promedio: {avg_score:.1f})")
    
    # Seleccionar portafolio DCA diversificado
    dca_portfolio = []
    
    # 1. Buscar QUALITY CHAMPION (40% del portafolio)
    if profile_groups['quality']:
        best_quality = profile_groups['quality'][0]
        dca_portfolio.append({
            'symbol': best_quality['symbol'],
            'allocation': 40,
            'role': '🏛️ CORE HOLDING',
            'profile': best_quality['profile_10y'],
            'score': best_quality['quality_score'],
            'rationale': 'Balance óptimo riesgo-retorno'
        })
    
    # 2. Buscar GROWTH component (35% del portafolio)
    growth_candidates = profile_groups['growth'] + profile_groups['consistent']
    if growth_candidates:
        # Evitar duplicar el core holding
        growth_candidates = [c for c in growth_candidates if c['symbol'] != dca_portfolio[0]['symbol']]
        if growth_candidates:
            best_growth = growth_candidates[0]
            dca_portfolio.append({
                'symbol': best_growth['symbol'],
                'allocation': 35,
                'role': '🚀 GROWTH ENGINE',
                'profile': best_growth['profile_10y'],
                'score': best_growth['quality_score'],
                'rationale': 'Motor de crecimiento del portafolio'
            })
    
    # 3. Buscar DEFENSIVE component (25% del portafolio)
    defensive_candidates = profile_groups['defensive'] + profile_groups['value']
    if defensive_candidates:
        # Evitar duplicados
        used_symbols = [p['symbol'] for p in dca_portfolio]
        defensive_candidates = [c for c in defensive_candidates if c['symbol'] not in used_symbols]
        if defensive_candidates:
            best_defensive = defensive_candidates[0]
            dca_portfolio.append({
                'symbol': best_defensive['symbol'],
                'allocation': 25,
                'role': '🛡️ STABILITY ANCHOR',
                'profile': best_defensive['profile_10y'],
                'score': best_defensive['quality_score'],
                'rationale': 'Estabilidad y protección en volatilidad'
            })
    
    # Mostrar recomendaciones
    if len(dca_portfolio) >= 2:
        print(f"\n💡 PORTAFOLIO DCA OPTIMIZADO POR PERFILES:")
        print("=" * 70)
        
        total_allocation = sum([p['allocation'] for p in dca_portfolio])
        for i, position in enumerate(dca_portfolio, 1):
            symbol = position['symbol']
            allocation = position['allocation']
            role = position['role']
            profile = position['profile']
            score = position['score']
            rationale = position['rationale']
            
            print(f"{i}. {symbol:6} ({allocation:2d}%) - {role}")
            print(f"   📊 Perfil: {profile}")
            print(f"   🎯 Score: {score:.1f}/100")
            print(f"   💭 Razón: {rationale}")
            print()
        
        # Implementación práctica
        print(f"📅 IMPLEMENTACIÓN MENSUAL ($1,000/mes):")
        monthly_amounts = []
        for position in dca_portfolio:
            amount = int(1000 * position['allocation'] / 100)
            monthly_amounts.append(amount)
            print(f"   • ${amount}/mes → {position['symbol']} (día 1 de cada mes)")
        
        # Proyección basada en scores
        weighted_score = sum([p['score'] * p['allocation']/100 for p in dca_portfolio])
        estimated_cagr = min(20, weighted_score * 0.25)  # Estimación conservadora
        
        print(f"\n📊 PROYECCIÓN 10 AÑOS:")
        print(f"   • Score ponderado del portafolio: {weighted_score:.1f}/100")
        print(f"   • CAGR estimado conservador: {estimated_cagr:.1f}%")
        
        monthly = 1000
        years = 10
        total_invested = monthly * 12 * years
        monthly_rate = estimated_cagr / 100 / 12
        
        if monthly_rate > 0:
            future_value = monthly * (((1 + monthly_rate) ** (12 * years) - 1) / monthly_rate)
            gain = future_value - total_invested
            print(f"   • Inversión total: ${total_invested:,.0f}")
            print(f"   • Valor proyectado: ${future_value:,.0f}")
            print(f"   • Ganancia potencial: ${gain:,.0f}")
        
        print(f"\n⚠️ REGLAS DCA CON PERFILES:")
        print("   • Rebalancear trimestralmente si alguna posición > 50%")
        print("   • Monitorear cambios de perfil (ej: Quality → Defensive)")
        print("   • En crashes >30%: acelerar DCA especialmente en QUALITY")
        print("   • Si un stock cambia a perfil negativo: considerar reemplazo")
        
        # Alternativas por perfil
        print(f"\n🔄 ALTERNATIVAS POR CATEGORÍA:")
        for profile_type, candidates in profile_groups.items():
            if len(candidates) > 1:
                top_3 = candidates[:3]
                symbols = [c['symbol'] for c in top_3]
                print(f"   {profile_type.upper()}: {', '.join(symbols)}")
    
    else:
        print("\n⚠️ No se encontraron suficientes candidatos para portafolio diversificado")
        print("Considera ETFs diversificados o relajar criterios")

# Agregar método mejorado a la clase
def _generate_dca_recommendations_method(self, all_results):
    """Método mejorado para generar recomendaciones DCA con perfiles"""
    _generate_dca_recommendations(self, all_results)

# Inyectar el método mejorado
EnhancedStockAnalyzer._generate_dca_recommendations = _generate_dca_recommendations_method

def enhanced_main():
    """Función principal mejorada"""
    try:
        print("🚀 STOCK ANALYZER - ARQUITECTURA OPTIMIZADA CON PERFILES")
        print("=" * 80)
        
        # Mostrar información de fechas dinámicas
        periods = DateManager.get_analysis_periods()
        print(f"\n📅 PERÍODOS DE ANÁLISIS DINÁMICOS:")
        for key, info in periods.items():
            print(f"   {info['name']}: {info['start_date']} → {info['end_date']}")
        
        # Crear analizador mejorado
        analyzer = EnhancedStockAnalyzer('Market_Cap_Ranked.xlsx')
        
        # Mostrar información del cache
        cache_status = analyzer.fetcher.cache_manager.get_cache_status_summary()
        print(f"\n{cache_status}")
        
        # Ejecutar análisis completo
        result = analyzer.run_complete_analysis(force_refresh=False)
        
        if result is None:
            print("❌ Análisis falló")
            return None
        
        filename, all_results, all_benchmarks, all_outperformers = result
        
        # Mostrar estadísticas finales
        print(f"\n📊 ESTADÍSTICAS FINALES:")
        for period_key in ['ytd', '5y', '10y']:
            if period_key in all_results:
                count = len(all_results[period_key])
                outperf_count = len(all_outperformers.get(period_key, []))
                period_name = periods[period_key]['name']
                print(f"   {period_name}: {count} acciones analizadas, {outperf_count} outperformers")
        
        # Mostrar algunos perfiles encontrados
        if '10y' in all_results and all_results['10y']:
            profile_counts = {}
            for result in all_results['10y']:
                profile = result.get('Investment_Profile', 'Unknown')
                profile_counts[profile] = profile_counts.get(profile, 0) + 1
            
            print(f"\n🏷️ PERFILES IDENTIFICADOS (10Y):")
            for profile, count in sorted(profile_counts.items(), key=lambda x: x[1], reverse=True)[:5]:
                print(f"   {profile}: {count} acciones")
        
        print(f"\n✅ ANÁLISIS COMPLETADO EXITOSAMENTE")
        print(f"📁 Archivo: {filename}")
        print(f"🔄 Próxima revisión: Trimestral")
        print(f"🎯 Nuevas funcionalidades: Perfiles automáticos + Score balanceado + DCA inteligente")
        
        return filename, all_results, all_benchmarks, all_outperformers
        
    except Exception as e:
        print(f"❌ Error en análisis: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    result = enhanced_main()